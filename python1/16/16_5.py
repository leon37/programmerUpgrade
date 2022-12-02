# 直“自愿”为“强制自愿俱乐部”记录会员会费。
# 这确实是一项枯燥的工作，包括维护一个电子表格，记录每个月谁交了会费，并用电子邮件提醒那些没交的会员。
# 不必你自己查看电子表格，而是向会费超期的会员复制和粘贴相同的电子邮件。
# 你猜对了，让我们编写一个脚本，帮你完成任务。
import sys
import openpyxl
import smtplib
import os
import datetime

os.chdir('python1\\automate_online-materials')
curDate = datetime.datetime.now().strftime('%b %Y')
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
latestMonth = sheet.cell(row=1, column=sheet.max_column).value
unpaidMembers = {}
for i in range(2, sheet.max_row+1):
    if sheet.cell(row=i, column=sheet.max_column).value != 'paid':
        name = sheet.cell(row=i, column=1).value
        email = sheet.cell(row=i, column=2).value
        unpaidMembers[name] = email

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('xxx@gmail.com', sys.argv[1])

for name, email in unpaidMembers.item():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not \
paid dues for %s. Please make this payment as soon as possible. Thank you!'" % \
        (latestMonth, name, latestMonth)
    sendmailStatus = smtpObj.sendmail('myemail@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()
