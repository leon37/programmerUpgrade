# 创建程序multiplicationTable.py，
# 从命令行接受数字N，在一个Excel电子表格中创建一个N×N的乘法表。
# 行1和列A应该用做标签，应该使用粗体。
#!python3
import openpyxl
import sys
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter

def genMultiTable(num):
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name('Sheet')
    for i in range(2, num+2):
        fontObj = Font(bold=True)
        styleObj = NamedStyle(font=fontObj, name='head')
        if i == 2:
            sheet['A'+str(i)].style=styleObj
            sheet[get_column_letter(i)+'1'].style=styleObj
        else:
            sheet['A'+str(i)].style='head'
            sheet[get_column_letter(i)+'1'].style='head'
        sheet['A'+str(i)] = i-1
        sheet[get_column_letter(i)+'1'] = i-1
    for i in range(2, num+2):
        for j in range (2, num+2):
            letter = get_column_letter(i)
            sheet[letter+str(j)].value = sheet[letter+'1'].value*sheet['A'+str(j)].value
    wb.save('multiplicationTable.xlsx')

if len(sys.argv) == 1:
    print('args not enough!')
else:
    param = sys.argv[1]
    genMultiTable(int(param))
