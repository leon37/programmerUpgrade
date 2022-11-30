import openpyxl
import os
import openpyxl.cell

os.chdir('python1\\12')
wb = openpyxl.load_workbook('example.xlsx')
# print(type(wb))
# print(wb.get_sheet_names())

# print(wb.get_sheet_by_name('Sheet3'))
sheet = wb.get_sheet_by_name('Sheet1')
# print(sheet['A1'].value)
# c = sheet['A1']
# print(c.coordinate)

# for i in range(1, 8, 2):
#     print(i, sheet.cell(row=i, column=2).value)

# print(sheet.get_highest_row())
# print(sheet.get_highest_column())

print(sheet.max_row)