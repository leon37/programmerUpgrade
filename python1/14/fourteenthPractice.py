# 为CSV文件，只要点几下鼠标，但如果有几百个Excel文件要转换为CSV，就需要点击几小时。
# 利用第12章的openpyxl模块，编程读取当前工作目录中的所有Excel文件，并输出为CSV文件。
# 一个Excel文件可能包含多个工作表，必须为每个表创建一个CSV文件。
# CSV文件的文件名应该是< Excel文件名>_< 表标题>.csv，
# 其中< Excel文件名>是没有扩展名的Excel文件名（例如'spam_data'，而不是'spam_data.xlsx'），
# <表标题>是Worksheet对象的title变量中的字符串。
import csv, openpyxl, os
from openpyxl.utils import get_column_letter

os.makedirs('csvSpreadsheets', exist_ok=True)
for excelFile in os.listdir('excelSpreadsheets'):
    wb = openpyxl.load_workbook(os.path.join('excelSpreadsheets', excelFile))
    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]
        contents = []
        for i in range(1, sheet.max_column):
            row = []
            for j in range(1, sheet.max_row):
                row.append(sheet[get_column_letter(j) + str(i)].value)
            contents.append(row)
        excelFileBaseName = os.path.splitext(excelFile)[0]
        newFilename = excelFileBaseName + '_' + sheetName + '.csv'
        newFile = open(os.path.join('csvSpreadsheets', newFilename), 'w', newline='')
        csvWriter = csv.writer(newFile)
        for row in contents:
            csvWriter.writerow(row)
        newFile.close()